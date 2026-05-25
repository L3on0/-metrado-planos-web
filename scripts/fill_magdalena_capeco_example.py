from pathlib import Path
import re
from collections import Counter

import fitz
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "data" / "templates" / "excel" / "capeco_metrado_templates.xlsx"
DOWNLOADS = Path(r"C:\Users\Mauro\Downloads")
OUT_DIR = DOWNLOADS / "METRADOS_MAGDALENA" / "02_metrados_generados"
OUT_FILE = OUT_DIR / "metrado_CAPECO_formateado_PLANO_MAGDALENA_1_PISO.xlsx"


def find_pdf():
    return next(DOWNLOADS.glob("PLANO MAGDALENA 1*PISO.pdf"))


def set_header(ws, pdf_name):
    values = {
        "B3": "Magdalena",
        "H3": "1 de 1",
        "B4": "Por definir",
        "H4": pdf_name,
        "B5": "2026-05-25",
        "H5": "Agente Codex",
        "B6": "Pendiente",
    }
    for cell, value in values.items():
        ws[cell] = value


def build():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    pdf = find_pdf()
    with fitz.open(pdf) as doc:
        text = doc[0].get_text("text") or ""

    tags = re.findall(r"\b([PVM]\d+)\*?\b", text)
    counts = Counter(tag[0] for tag in set(tags))
    upper = text.upper()
    finishes = {
        "Piso porcelanato detectado en plano": upper.count("PISO PORCELANATO"),
        "Piso cemento pulido detectado en plano": upper.count("PISO CEMENTO PULIDO"),
        "Falso cielorraso con baldosas de yeso detectado": upper.count("FCR BALDOSAS YESO"),
    }

    wb = load_workbook(TEMPLATE)
    ws = wb["CAPECO General"]
    set_header(ws, pdf.name)

    rows = [
        ("OE.3-VANOS-P", "Puertas identificadas por etiquetas P en plano", counts.get("P", 0), None, None, None, counts.get("P", 0), None, None, None, None, None, counts.get("P", 0), "und"),
        ("OE.3-VANOS-V", "Ventanas identificadas por etiquetas V en plano", counts.get("V", 0), None, None, None, counts.get("V", 0), None, None, None, None, None, counts.get("V", 0), "und"),
        ("OE.3-EQ-M", "Elementos identificados con etiqueta M, pendiente de leyenda", counts.get("M", 0), None, None, None, counts.get("M", 0), None, None, None, None, None, counts.get("M", 0), "und"),
        ("OE.3.4.2", "Piso porcelanato. Area pendiente de asociar por ambiente", finishes["Piso porcelanato detectado en plano"], None, None, None, None, None, None, None, None, None, None, "m2"),
        ("OE.3.4.2", "Piso cemento pulido. Area pendiente de asociar por ambiente", finishes["Piso cemento pulido detectado en plano"], None, None, None, None, None, None, None, None, None, None, "m2"),
        ("OE.3.3.6", "Falso cielorraso baldosas yeso. Area pendiente", finishes["Falso cielorraso con baldosas de yeso detectado"], None, None, None, None, None, None, None, None, None, None, "m2"),
    ]

    start = 11
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = value

    resumen = wb["CAPECO Revision Parametrica"]
    set_header(resumen, pdf.name)
    resumen["A11"] = "Puertas detectadas"
    resumen["B11"] = counts.get("P", 0)
    resumen["F11"] = "Conteo preliminar por etiquetas P; validar con cuadro de vanos."
    resumen["A12"] = "Ventanas detectadas"
    resumen["B12"] = counts.get("V", 0)
    resumen["F12"] = "Conteo preliminar por etiquetas V; validar con cuadro de vanos."
    resumen["A13"] = "Acabados en m2"
    resumen["F13"] = "Pendiente de calculo de areas por ambiente desde geometria/cotas."

    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    build()
