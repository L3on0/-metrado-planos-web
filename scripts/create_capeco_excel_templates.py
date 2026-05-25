from pathlib import Path
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "templates" / "excel"
OUT_FILE = OUT_DIR / "capeco_metrado_templates.xlsx"


THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
DOUBLE = Side(style="double", color="000000")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
BORDER_BOTTOM_DOUBLE = Border(bottom=DOUBLE)

FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_HEADER = Font(name="Arial", size=9, bold=True)
FONT_BODY = Font(name="Arial", size=9)
FONT_SMALL = Font(name="Arial", size=8)

FILL_HEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_SUBHEADER = PatternFill("solid", fgColor="EAF3F8")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")


def set_page(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.freeze_panes = "A12"


def style_range(ws, cell_range, border=BORDER_THIN, font=FONT_BODY, fill=None, align=None):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = copy(border)
            cell.font = copy(font)
            if fill:
                cell.fill = copy(fill)
            if align:
                cell.alignment = copy(align)


def merge_value(ws, range_name, value, font=FONT_BODY, fill=None, border=BORDER_THIN, align=None):
    ws.merge_cells(range_name)
    cell = ws[range_name.split(":")[0]]
    cell.value = value
    cell.font = font
    cell.border = copy(border)
    if fill:
        cell.fill = copy(fill)
    target_align = align or Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.alignment = copy(target_align)
    style_range(ws, range_name, border=border, font=font, fill=fill, align=target_align)


def apply_header_block(ws, last_col):
    merge_value(ws, f"A1:{last_col}1", "METRADO", FONT_TITLE, align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 24

    labels = [
        ("A3", "Obra", "B3:F3"),
        ("G3", "Hoja Nro.", "H3:J3"),
        ("A4", "Propietario", "B4:F4"),
        ("G4", "Plano Nro.", "H4:J4"),
        ("A5", "Fecha", "B5:F5"),
        ("G5", "Hecho por", "H5:J5"),
        ("A6", "Revisado", "B6:F6"),
    ]
    for label_cell, label, value_range in labels:
        ws[label_cell] = label
        ws[label_cell].font = FONT_HEADER
        ws[label_cell].alignment = Alignment(horizontal="right", vertical="center")
        ws[label_cell].border = BORDER_THIN
        merge_value(ws, value_range, "", FONT_BODY, FILL_INPUT, align=Alignment(horizontal="left", vertical="center"))

    for row in range(3, 7):
        ws.row_dimensions[row].height = 18


def setup_widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def capeco_general(wb):
    ws = wb.create_sheet("CAPECO General")
    set_page(ws)
    widths = [9, 28, 10, 9, 9, 9, 10, 10, 9, 9, 9, 10, 10, 7]
    setup_widths(ws, widths)
    apply_header_block(ws, "N")

    merge_value(ws, "A8:A10", "Partida\nNro.", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "B8:B10", "Especificaciones", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "C8:C10", "Nro. de\nveces", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "D8:F8", "Medidas", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    ws["D9"] = "Largo"
    ws["E9"] = "Ancho"
    ws["F9"] = "Altura"
    merge_value(ws, "G8:G10", "Parcial", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "H8:L8", "Vanos o disminuciones", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    ws["H9"] = "Nro. de\nveces"
    ws["I9"] = "Largo"
    ws["J9"] = "Ancho"
    ws["K9"] = "Altura"
    ws["L9"] = "Parcial"
    merge_value(ws, "M8:M10", "Total", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "N8:N10", "Und", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)

    style_range(ws, "D9:F10", BORDER_MEDIUM, FONT_HEADER, FILL_HEADER, Alignment(horizontal="center", vertical="center", wrap_text=True))
    style_range(ws, "H9:L10", BORDER_MEDIUM, FONT_HEADER, FILL_HEADER, Alignment(horizontal="center", vertical="center", wrap_text=True))

    for row in range(11, 41):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER_THIN
            cell.font = FONT_BODY
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=7).value = f'=IF(C{row}="","",C{row}*D{row}*E{row}*F{row})'
        ws.cell(row=row, column=12).value = f'=IF(H{row}="","",H{row}*I{row}*J{row}*K{row})'
        ws.cell(row=row, column=13).value = f'=IF(G{row}="","",G{row}-IF(L{row}="",0,L{row}))'
        ws.row_dimensions[row].height = 22

    ws["A42"] = "TOTAL"
    ws["A42"].font = FONT_HEADER
    ws["A42"].alignment = Alignment(horizontal="right")
    ws.merge_cells("A42:L42")
    ws["M42"] = "=SUM(M11:M40)"
    ws["N42"] = ""
    style_range(ws, "A42:N42", BORDER_MEDIUM, FONT_HEADER, FILL_SUBHEADER, Alignment(horizontal="center", vertical="center"))
    return ws


def capeco_concreto_armado(wb):
    ws = wb.create_sheet("CAPECO Concreto Armado")
    set_page(ws)
    widths = [9, 20, 24, 10, 9, 9, 9, 12, 12, 12]
    setup_widths(ws, widths)
    apply_header_block(ws, "J")
    merge_value(ws, "A8:A10", "Partida\nNro.", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "B8:B10", "Elemento", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "C8:C10", "Descripcion", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "D8:D10", "Cant. de\nelementos", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "E8:G8", "Medidas", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    ws["E9"] = "Largo"
    ws["F9"] = "Ancho"
    ws["G9"] = "Altura"
    merge_value(ws, "H8:H10", "Concreto\nTotal (m3)", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "I8:I10", "Encofrado\nTotal (m2)", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    merge_value(ws, "J8:J10", "Fierro\nTotal (kg)", FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    style_range(ws, "E9:G10", BORDER_MEDIUM, FONT_HEADER, FILL_HEADER, Alignment(horizontal="center", vertical="center", wrap_text=True))

    for row in range(11, 41):
        for col in range(1, 11):
            c = ws.cell(row=row, column=col)
            c.border = BORDER_THIN
            c.font = FONT_BODY
            c.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=8).value = f'=IF(D{row}="","",D{row}*E{row}*F{row}*G{row})'
        ws.row_dimensions[row].height = 22
    style_range(ws, "A42:J42", BORDER_MEDIUM, FONT_HEADER, FILL_SUBHEADER, Alignment(horizontal="center", vertical="center"))
    ws.merge_cells("A42:G42")
    ws["A42"] = "TOTAL"
    ws["H42"] = "=SUM(H11:H40)"
    ws["I42"] = "=SUM(I11:I40)"
    ws["J42"] = "=SUM(J11:J40)"
    return ws


def capeco_fierro(wb):
    ws = wb.create_sheet("CAPECO Fierro")
    set_page(ws)
    widths = [9, 26, 18, 10, 9, 12, 14, 12, 12]
    setup_widths(ws, widths)
    apply_header_block(ws, "I")
    headers = [
        ("A8:A10", "Partida\nNro."),
        ("B8:B10", "Descripcion"),
        ("C8:C10", "Diseno del\nfierro"),
        ("D8:D10", "Diametro"),
        ("E8:E10", "Cant."),
        ("F8:F10", "Longitud"),
        ("G8:G10", "Longitud\nTotal"),
        ("H8:H10", "Peso\nkg/ml"),
        ("I8:I10", "Total\n(kg)")
    ]
    for rng, value in headers:
        merge_value(ws, rng, value, FONT_HEADER, FILL_HEADER, BORDER_MEDIUM)
    for row in range(11, 41):
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.border = BORDER_THIN
            c.font = FONT_BODY
            c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=7).value = f'=IF(E{row}="","",E{row}*F{row})'
        ws.cell(row=row, column=9).value = f'=IF(G{row}="","",G{row}*H{row})'
        ws.row_dimensions[row].height = 22
    style_range(ws, "A42:I42", BORDER_MEDIUM, FONT_HEADER, FILL_SUBHEADER, Alignment(horizontal="center", vertical="center"))
    ws.merge_cells("A42:H42")
    ws["A42"] = "TOTAL"
    ws["I42"] = "=SUM(I11:I40)"
    return ws


def capeco_parametrica(wb):
    ws = wb.create_sheet("CAPECO Revision Parametrica")
    set_page(ws)
    widths = [28, 18, 18, 18, 14, 32]
    setup_widths(ws, widths)
    apply_header_block(ws, "F")
    merge_value(ws, "A8:F8", "CUADRO DE COMPARACION PARAMETRICA DE METRADOS", FONT_TITLE, align=Alignment(horizontal="center", vertical="center"))
    headers = ["Concepto", "Metrado base", "Metrado comparado", "Parametro de obra", "Rango", "Observacion"]
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=10, column=col)
        c.value = header
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER_MEDIUM
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(11, 31):
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.border = BORDER_THIN
            c.font = FONT_BODY
            c.alignment = Alignment(horizontal="center" if col != 1 and col != 6 else "left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=4).value = f'=IF(B{row}="","",B{row}/C{row})'
        ws.row_dimensions[row].height = 24
    return ws


def build():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    capeco_general(wb)
    capeco_concreto_armado(wb)
    capeco_fierro(wb)
    capeco_parametrica(wb)
    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    build()
