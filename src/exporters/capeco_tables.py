"""
Tablas CAPECO para el pipeline legacy de PDF analysis.

Contiene la lógica hardcodeada del proyecto "Magdalena" que analiza
texto de PDF buscando etiquetas P, V, M y acabados específicos.

NOTA: Este módulo es el flujo LEGACY. Los nuevos procesadores (DXF/DWG/PDF)
usan el pipeline estándar: extract_measurements() → build_metrado() → build_excel().
"""
from __future__ import annotations

from collections import Counter
from io import BytesIO
from pathlib import Path
import re

import fitz
import pandas as pd
from openpyxl import load_workbook

from src.log_utils import get_logger

logger = get_logger(__name__)

CAPECO_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "data" / "templates" / "excel" / "capeco_metrado_templates.xlsx"
)

# Ambientes predefinidos para el proyecto "Magdalena"
_AMBIENTES_BASE = [
    "IMAGEN INSTITUCIONAL", "SALA DE ESTAR", "SECRETARÍA", "COMEDOR",
    "DATA CENTER", "SS.HH. 1", "PASILLO 1", "PASADIZO",
    "SS.HH. DISC.", "HALL 2", "HALL 1", "SALA DE REUNIONES",
    "SSHH 2", "DUCTO", "Rampa",
]


def analyze_pdf_text(file_path: Path) -> dict:
    """Analiza el texto de un PDF en busca de etiquetas P, V, M, acabados y ambientes.

    Args:
        file_path: Ruta al archivo PDF.

    Returns:
        Dict con texto extraído, etiquetas, acabados y ambientes detectados.
    """
    with fitz.open(file_path) as doc:
        page = doc[0]
        text = page.get_text("text") or ""
        drawings = page.get_drawings()
        rect = page.rect

    tags = re.findall(r"\b([PVM]\d+)\*?\b", text)
    unique_tags = sorted(
        set(tags),
        key=lambda tag: (tag[0], int(re.search(r"\d+", tag).group())),
    )
    counts = Counter(tag[0] for tag in unique_tags)
    upper = text.upper()

    finishes = {
        "PISO PORCELANATO": upper.count("PISO PORCELANATO"),
        "PISO CEMENTO PULIDO": upper.count("PISO CEMENTO PULIDO"),
        "FCR BALDOSAS YESO": upper.count("FCR BALDOSAS YESO"),
    }
    ambientes = [name for name in _AMBIENTES_BASE if name.upper() in upper]

    return {
        "page_width": rect.width,
        "page_height": rect.height,
        "text": text,
        "text_chars": len(text),
        "drawings_count": len(drawings),
        "tags": unique_tags,
        "counts": counts,
        "finishes": finishes,
        "ambientes": ambientes,
    }


def build_capeco_tables(file_name: str, analysis: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Construye tablas CAPECO hardcodeadas a partir del análisis de texto.

    Args:
        file_name: Nombre del archivo.
        analysis: Dict devuelto por analyze_pdf_text().

    Returns:
        (metrado_df, resumen_df, diag_df)
    """
    counts = analysis["counts"]
    finishes = analysis["finishes"]

    rows = [
        {
            "Partida Nro.": "OE.3-VANOS-P",
            "Especificaciones": "Puertas identificadas por etiquetas P en plano",
            "Nro. de veces": counts.get("P", 0),
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": counts.get("P", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("P", 0),
            "Und": "und",
            "Nota de revision": "Preliminar. Falta contrastar con cuadro de vanos si existe.",
        },
        {
            "Partida Nro.": "OE.3-VANOS-V",
            "Especificaciones": "Ventanas identificadas por etiquetas V en plano",
            "Nro. de veces": counts.get("V", 0),
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": counts.get("V", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("V", 0),
            "Und": "und",
            "Nota de revision": "Preliminar. Falta contrastar con cuadro de vanos si existe.",
        },
        {
            "Partida Nro.": "OE.3-EQ-M",
            "Especificaciones": "Elementos identificados con etiqueta M",
            "Nro. de veces": counts.get("M", 0),
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": counts.get("M", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("M", 0),
            "Und": "und",
            "Nota de revision": "Requiere leyenda del plano para definir si es mampara, mueble u otro elemento.",
        },
        {
            "Partida Nro.": "OE.3.4.2",
            "Especificaciones": "Piso porcelanato detectado en ambientes del primer nivel",
            "Nro. de veces": finishes["PISO PORCELANATO"],
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None, "Und": "m2",
            "Nota de revision": "Cantidad pendiente: falta delimitar areas por ambiente o leer DWG por capas.",
        },
        {
            "Partida Nro.": "OE.3.4.2",
            "Especificaciones": "Piso cemento pulido detectado",
            "Nro. de veces": finishes["PISO CEMENTO PULIDO"],
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None, "Und": "m2",
            "Nota de revision": "Cantidad pendiente: falta delimitar area del ambiente correspondiente.",
        },
        {
            "Partida Nro.": "OE.3.3.6",
            "Especificaciones": "Falso cielorraso con baldosas de yeso detectado",
            "Nro. de veces": finishes["FCR BALDOSAS YESO"],
            "Largo": None, "Ancho": None, "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None, "Und": "m2",
            "Nota de revision": "Cantidad pendiente: requiere asociar poligono de cielorraso por ambiente.",
        },
    ]
    metrado_df = pd.DataFrame(rows)

    resumen_df = pd.DataFrame([
        {"Concepto": "Puertas", "Partida": "OE.3-VANOS-P", "Und": "und",
         "Total": counts.get("P", 0), "Estado": "Preliminar por conteo de etiquetas"},
        {"Concepto": "Ventanas", "Partida": "OE.3-VANOS-V", "Und": "und",
         "Total": counts.get("V", 0), "Estado": "Preliminar por conteo de etiquetas"},
        {"Concepto": "Elementos M", "Partida": "OE.3-EQ-M", "Und": "und",
         "Total": counts.get("M", 0), "Estado": "Pendiente de leyenda"},
        {"Concepto": "Piso porcelanato", "Partida": "OE.3.4.2", "Und": "m2",
         "Total": None, "Estado": "Pendiente de area"},
        {"Concepto": "Piso cemento pulido", "Partida": "OE.3.4.2", "Und": "m2",
         "Total": None, "Estado": "Pendiente de area"},
        {"Concepto": "Falso cielorraso baldosas yeso", "Partida": "OE.3.3.6", "Und": "m2",
         "Total": None, "Estado": "Pendiente de area"},
    ])
    diag_df = pd.DataFrame([
        {"Campo": "Archivo", "Valor": file_name},
        {"Campo": "Especialidad estimada", "Valor": "Arquitectura"},
        {"Campo": "Formato de salida", "Valor": "CAPECO General"},
        {"Campo": "Tamaño pagina PDF", "Valor": f"{analysis['page_width']:.1f} x {analysis['page_height']:.1f}"},
        {"Campo": "Texto extraido", "Valor": analysis["text_chars"]},
        {"Campo": "Objetos vectoriales", "Valor": analysis["drawings_count"]},
        {"Campo": "Etiquetas detectadas", "Valor": ", ".join(analysis["tags"])},
        {"Campo": "Ambientes detectados", "Valor": ", ".join(analysis["ambientes"])},
    ])
    return metrado_df, resumen_df, diag_df


def build_formatted_capeco_excel(file_name: str, metrado_df: pd.DataFrame,
                                 resumen_df: pd.DataFrame) -> bytes:
    """Genera Excel con formato CAPECO a partir de tablas existentes.

    Args:
        file_name: Nombre del archivo.
        metrado_df: DataFrame del metrado.
        resumen_df: DataFrame del resumen.

    Returns:
        Bytes del archivo .xlsx.
    """
    if not CAPECO_TEMPLATE.exists():
        logger.warning(f"Plantilla CAPECO no encontrada en {CAPECO_TEMPLATE}")
        return b""

    wb = load_workbook(CAPECO_TEMPLATE)
    ws = wb["CAPECO General"]
    header_values = {
        "B3": "Magdalena",
        "H3": "1 de 1",
        "B4": "Por definir",
        "H4": file_name,
        "B5": "2026-05-25",
        "H5": "SAS metrados",
        "B6": "Pendiente",
    }
    for cell, value in header_values.items():
        ws[cell] = value

    template_cols = [
        "Partida Nro.", "Especificaciones", "Nro. de veces",
        "Largo", "Ancho", "Altura", "Parcial",
        "Vanos o disminuciones - Nro. veces",
        "Vanos o disminuciones - Largo",
        "Vanos o disminuciones - Ancho",
        "Vanos o disminuciones - Altura",
        "Vanos o disminuciones - Parcial",
        "Total", "Und",
    ]
    for row_idx, row in enumerate(
        metrado_df[template_cols].itertuples(index=False, name=None), start=11
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = None if pd.isna(value) else value

    rev = wb["CAPECO Revision Parametrica"]
    for cell, value in header_values.items():
        rev[cell] = value
    for row_idx, row in enumerate(resumen_df.itertuples(index=False), start=11):
        rev.cell(row=row_idx, column=1).value = row.Concepto
        rev.cell(row=row_idx, column=2).value = row.Total
        rev.cell(row=row_idx, column=6).value = row.Estado

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
