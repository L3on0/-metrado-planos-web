"""
Formato de metrado de Arquitectura según RD-073-2010-VIVIENDA-VMCS-DNC.

Cada elemento contable (puertas, ventanas, columnas, etc.) se lista
individualmente con su ubicación referenciada a ejes del plano o por
ambiente, en lugar de agruparse en un total global.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.classification import classify_layer
from src.measurements import Measurement, MetradoItem

# ---------------------------------------------------------------------------
# Datos
# ---------------------------------------------------------------------------

# Contadores secuenciales por tipo de elemento
_ITEM_COUNTERS: dict[str, int] = defaultdict(int)

_PREFIX_MAP = {
    "ARQ-05": "P",   # Puertas
    "ARQ-06": "V",   # Ventanas
    "ARQ-08": "BA",  # Barandas
    "EST-01": "C",   # Columnas
    "EST-02": "VG",  # Vigas
    "ISS-03": "AS",  # Aparatos sanitarios
    "ISE-02": "T",   # Tomacorrientes
    "ISE-04": "TE",  # Tableros
}


def _reset_counters():
    _ITEM_COUNTERS.clear()


def _next_id(partida: str) -> str:
    prefix = _PREFIX_MAP.get(partida, "XX")
    _ITEM_COUNTERS[prefix] += 1
    return f"{prefix}-{_ITEM_COUNTERS[prefix]:02d}"


def _infer_location(m: Measurement, axes: Optional[list] = None) -> str:
    """Infere la ubicación de un elemento según coordenadas y ejes del plano.

    Args:
        m: Measurement con coordenadas (coord_x, coord_y).
        axes: Lista de ejes detectados [(label, x, y, orientation), ...].

    Returns:
        String con la ubicación (ej: "Eje B/2" o nombre de capa).
    """
    # Intentar ubicar por ejes
    if axes and (m.coord_x != 0.0 or m.coord_y != 0.0):
        return _nearest_axis_label(m.coord_x, m.coord_y, axes)

    # Fallback: extraer nombre descriptivo de la capa
    layer = m.layer
    # Limpiar prefijos comunes
    for prefix in ["0. ", "ARQ ", "ARQ-", "EST ", "EST-"]:
        if prefix in layer:
            parts = layer.split(prefix)
            if len(parts) > 1:
                return parts[-1].strip()
    return layer


# ---------------------------------------------------------------------------
# Detección de ejes
# ---------------------------------------------------------------------------


def extract_axes(all_measurements: list[Measurement]) -> list[tuple]:
    """Detecta ejes del plano a partir de mediciones en capas de referencia.

    Busca líneas en capas clasificadas como REF-02 (Ejes) o líneas muy
    largas que cruzan el dibujo (típicas de ejes).

    Returns:
        Lista de (label, x, y, orientation) donde orientation es 'V' (vertical)
        u 'H' (horizontal). label es un número (1,2,3) o letra (A,B,C).
    """
    from src.classification import classify_layer

    axis_data: list[tuple] = []

    for m in all_measurements:
        rule = classify_layer(m.layer)
        es_eje = rule.partida == "REF-02"
        es_linea_larga = (
            m.element_type in ("line", "polyline")
            and m.quantity > 5.0  # líneas largas = probables ejes
            and rule.partida != "REF-01"  # no cotas
        )

        if not (es_eje or es_linea_larga):
            continue

        # Determinar orientación por diferencia de coordenadas
        if hasattr(m, 'coord_x') and hasattr(m, 'coord_y'):
            orient = "V"  # asumir vertical
            # Si el elemento tiene coordenadas de inicio y fin, calcular
            # orientación real. Por ahora usar una heurística simple:
            # las líneas verticales tienen x constante, horizontales y constante
            # estimar por diferencia
            dx = abs(m.coord_x) if m.coord_x else 0
            dy = abs(m.coord_y) if m.coord_y else 0
            orient = "V" if dx > dy else "H"

            axis_data.append((m.layer, m.coord_x, m.coord_y, orient))

    if not axis_data:
        return []

    # Agrupar ejes por orientación y ordenar
    verticales = sorted(
        [(label, x, y) for label, x, y, o in axis_data if o == "V"],
        key=lambda t: t[1],  # ordenar por x
    )
    horizontales = sorted(
        [(label, y, x) for label, x, y, o in axis_data if o == "H"],
        key=lambda t: t[1],  # ordenar por y
    )

    # Asignar etiquetas: números a verticales, letras a horizontales
    result: list[tuple] = []
    for i, (label, x, y) in enumerate(verticales):
        result.append((str(i + 1), x, y, "V"))
    for i, (label, y_coord, _) in enumerate(horizontales):
        result.append((chr(65 + i), 0, y_coord, "H"))  # A, B, C...

    return result


def _nearest_axis_label(x: float, y: float, axes: list[tuple]) -> str:
    """Encuentra el eje vertical y horizontal más cercano a un punto.

    Args:
        x, y: Coordenadas del elemento.
        axes: Lista de (label, x_axis, y_axis, orientation).

    Returns:
        String "Eje {horizontal}/{vertical}" (ej: "Eje B/2").
    """
    nearest_v: tuple[float, str] = (float("inf"), "?")
    nearest_h: tuple[float, str] = (float("inf"), "?")

    for label, ax, ay, orient in axes:
        if orient == "V":
            dist = abs(x - ax)
            if dist < nearest_v[0]:
                nearest_v = (dist, label)
        else:
            dist = abs(y - ay)
            if dist < nearest_h[0]:
                nearest_h = (dist, label)

    near_v_label = nearest_v[1] if nearest_v[1] != "?" else ""
    near_h_label = nearest_h[1] if nearest_h[1] != "?" else ""

    near_v_dist = nearest_v[0]
    near_h_dist = nearest_h[0]

    # Si no hay ejes cercanos, devolver coordenadas planas
    if near_v_dist > 50 and near_h_dist > 50:
        return f"~({x:.1f}, {y:.1f})"
    if near_v_dist > 50:
        return f"Eje {near_h_label}"
    if near_h_dist > 50:
        return f"Eje {near_v_label}"

    return f"Eje {near_h_label}/{near_v_label}"


# ---------------------------------------------------------------------------
# Item detallado de Arquitectura
# ---------------------------------------------------------------------------

@dataclass
class ItemArquitectura:
    """Representa un elemento individual en el formato de Arquitectura."""
    item_id: str
    partida: str
    ubicacion: str
    descripcion: str
    largo: Optional[float] = None
    ancho: Optional[float] = None
    alto: Optional[float] = None
    unidad: str = "und"
    cantidad: float = 1.0
    observacion: str = ""


def measurements_to_arquitectura(measurements: list[Measurement]) -> list[ItemArquitectura]:
    """Convierte mediciones crudas en items individuales de Arquitectura.

    Solo procesa elementos contables (und) y los lista uno por uno.
    Los elementos medibles (m, m2) se agrupan normalmente.
    """
    _reset_counters()
    items: list[ItemArquitectura] = []

    # Detectar ejes del plano
    axes = extract_axes(measurements)

    for m in measurements:
        rule = classify_layer(m.layer)
        partida = rule.partida
        es_contable = partida in {
            "ARQ-05", "ARQ-06", "ARQ-08",
            "EST-01", "EST-02",
            "ISS-03", "ISE-02", "ISE-04",
        }

        if es_contable:
            item_id = _next_id(partida)
            ubicacion = _infer_location(m, axes)
            # Determinar dimensiones aproximadas si están disponibles
            largo = m.quantity if m.unit == "m" else None
            ancho = None
            alto = None

            items.append(ItemArquitectura(
                item_id=item_id,
                partida=partida,
                ubicacion=ubicacion,
                descripcion=m.description or rule.descripcion,
                largo=largo,
                ancho=ancho,
                alto=alto,
                unidad="und",
                cantidad=1.0,
                observacion=f"Extraído de capa: {m.layer}",
            ))

    return items


def metrado_to_arquitectura(items: list[MetradoItem]) -> list[ItemArquitectura]:
    """Convierte MetradoItems agrupados en items individuales de Arquitectura.

    Útil cuando ya se tiene el metrado agrupado pero se quiere
    desglosar visualmente en el formato de Arquitectura.
    """
    _reset_counters()
    result: list[ItemArquitectura] = []

    for item in items:
        if item.unidad != "und":
            continue  # solo und se desglosa
        count = int(item.cantidad)
        for i in range(count):
            item_id = _next_id(item.partida)
            result.append(ItemArquitectura(
                item_id=item_id,
                partida=item.partida,
                ubicacion="—",
                descripcion=item.descripcion,
                unidad="und",
                cantidad=1.0,
                observacion=f"Item {i+1} de {count}" if count > 1 else "",
            ))

    return result


# ---------------------------------------------------------------------------
# Exportación Excel
# ---------------------------------------------------------------------------

# Estilos
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1f2937")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="ffffff")
_DATA_FONT = Font(name="Calibri", size=9, color="1f2937")
_SMALL_FONT = Font(name="Calibri", size=8, italic=True, color="6b7280")
_HEADER_FILL = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
_PUERTAS_FILL = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")  # amarillo claro
_VENTANAS_FILL = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")  # azul claro
_ESTRUCT_FILL = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")  # verde claro
_THIN_BORDER = Border(
    left=Side(style="thin", color="d1d5db"),
    right=Side(style="thin", color="d1d5db"),
    top=Side(style="thin", color="d1d5db"),
    bottom=Side(style="thin", color="d1d5db"),
)


def _fill_for_partida(partida: str) -> PatternFill:
    if partida.startswith("ARQ-05"):
        return _PUERTAS_FILL
    if partida.startswith("ARQ-06"):
        return _VENTANAS_FILL
    if partida.startswith("EST-"):
        return _ESTRUCT_FILL
    return PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")


def build_arquitectura_excel(items: list[ItemArquitectura],
                              proyecto: str = "Sin proyecto",
                              especialidad: str = "Arquitectura") -> bytes:
    """Genera un Excel con formato de metrado de Arquitectura.

    Cada elemento se lista individualmente con:
        Item | Ubicación (Eje/Ambiente) | Descripción | Largo | Ancho | Alto | Und | Cant.
    """
    wb = __import__("openpyxl").Workbook()
    ws = wb.active
    ws.title = "Metrado Arquitectura"

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"METRADO DE ARQUITECTURA - {proyecto}"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # Subtítulo
    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        f"Según RD-073-2010-VIVIENDA-VMCS-DNC | "
        f"Items referenciados por ubicación | {especialidad}"
    )
    ws["A2"].font = _SMALL_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = ["Item", "Partida", "Ubicación (Eje/Ambiente)", "Descripción",
               "Largo (m)", "Ancho (m)", "Alto (m)", "Und", "Cant."]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    # Anchos de columna
    widths = [8, 10, 28, 40, 10, 10, 10, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Datos
    for i, item in enumerate(items):
        row = 5 + i
        fill = _fill_for_partida(item.partida)

        values = [
            item.item_id,
            item.partida,
            item.ubicacion,
            item.descripcion,
            item.largo if item.largo else "",
            item.ancho if item.ancho else "",
            item.alto if item.alto else "",
            item.unidad,
            item.cantidad,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.font = _DATA_FONT
            cell.fill = fill
            cell.border = _THIN_BORDER
            if col > 1 and isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            elif col <= 3:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # Altura de fila
        ws.row_dimensions[row].height = 18

    # Resumen al final
    row_resumen = 5 + len(items) + 2
    ws.merge_cells(f"A{row_resumen}:C{row_resumen}")
    ws.cell(row=row_resumen, column=1).value = "RESUMEN"
    ws.cell(row=row_resumen, column=1).font = Font(name="Calibri", size=11, bold=True)

    from collections import Counter
    partida_counts = Counter(it.partida for it in items)
    for i, (partida, count) in enumerate(partida_counts.most_common()):
        r = row_resumen + 1 + i
        ws.cell(row=r, column=1).value = partida
        ws.cell(row=r, column=2).value = f"{count} elementos"
        ws.cell(row=r, column=2).font = _DATA_FONT

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def build_arquitectura_pdf(items: list[ItemArquitectura],
                            proyecto: str = "Sin proyecto") -> bytes:
    """Genera un PDF con formato de metrado de Arquitectura."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                     Spacer, Table, TableStyle)

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    story = []
    style_title = ParagraphStyle("T", fontSize=14, bold=True, alignment=1)
    style_sub = ParagraphStyle("S", fontSize=8, alignment=1,
                                textColor=rl_colors.HexColor("#6b7280"))
    style_data = ParagraphStyle("D", fontSize=7, leading=9)

    story.append(Paragraph(f"METRADO DE ARQUITECTURA", style_title))
    story.append(Paragraph(f"{proyecto} | RD-073-2010-VIVIENDA-VMCS-DNC", style_sub))
    story.append(Spacer(1, 8))

    # Tabla
    header = ["Item", "Partida", "Ubicación", "Descripción", "Largo", "Ancho", "Alto", "Und", "Cant."]
    data = [header]
    for it in items:
        data.append([
            it.item_id, it.partida, it.ubicacion,
            Paragraph(it.descripcion, style_data),
            f"{it.largo:.2f}" if it.largo else "",
            f"{it.ancho:.2f}" if it.ancho else "",
            f"{it.alto:.2f}" if it.alto else "",
            it.unidad, str(it.cantidad),
        ])

    col_w = [1.5*cm, 1.5*cm, 3.5*cm, 5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.2*cm, 1.2*cm]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    story.append(tbl)

    doc.build(story)
    output.seek(0)
    return output.read()
