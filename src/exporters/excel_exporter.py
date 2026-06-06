"""
Exportador profesional de metrados a Excel con formato CAPECO.

Genera un libro con hojas de aspecto profesional, listo para entregar
a un cliente o contratista.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.measurements import MetradoItem

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1f2937")
_SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="374151")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="ffffff")
_DATA_FONT = Font(name="Calibri", size=10, color="1f2937")
_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="374151")
_VALUE_FONT = Font(name="Calibri", size=10, color="1f2937")
_NOTE_FONT = Font(name="Calibri", size=8, italic=True, color="6b7280")

_HEADER_FILL = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
_ROW_EVEN_FILL = PatternFill(start_color="f9fafb", end_color="f9fafb", fill_type="solid")
_ROW_ODD_FILL = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
_INFO_LABEL_FILL = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin", color="d1d5db"),
    right=Side(style="thin", color="d1d5db"),
    top=Side(style="thin", color="d1d5db"),
    bottom=Side(style="thin", color="d1d5db"),
)

_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")
_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Datos del ingeniero (edítalos con tus datos reales)
# ---------------------------------------------------------------------------
INGENIERO = {
    "nombre": "Mauro Cruz Balladares",
    "cip": "130410",
    "dni": "43151426",
    "telefono": "987547728",
    "email": "mauro.cruz.balladares@gmail.com",
    "direccion": "Mz. G Lt 7 Urb. Virgen del Sol, Los Olivos, Lima",
    "ruc": "10431514261",
}


# ---------------------------------------------------------------------------
# Construcción del Excel
# ---------------------------------------------------------------------------

def _apply_cell(cell, value, font=_DATA_FONT, fill=None, alignment=_LEFT_ALIGNMENT,
                border=_THIN_BORDER, number_format=None):
    """Aplica estilo a una celda y asigna su valor."""
    cell.value = value
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if number_format:
        cell.number_format = number_format


def _write_header_block(ws, proyecto: str, especialidad: str, fecha: Optional[date] = None,
                        escala: float = 1.0):
    """Escribe el bloque de encabezado con datos del ingeniero y proyecto."""
    today = fecha or date.today()

    # Título
    ws.merge_cells("A1:N1")
    _apply_cell(ws["A1"], "METRADO DE PLANOS", _TITLE_FONT,
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # Separador
    ws.merge_cells("A2:N2")
    _apply_cell(ws["A2"], "", font=_DATA_FONT,
                fill=PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid"))
    ws.row_dimensions[2].height = 4

    # Bloque de información: Ingeniero (izquierda) + Proyecto (derecha)
    info_data = [
        ("Ingeniero:", INGENIERO["nombre"]),
        ("CIP:", INGENIERO["cip"]),
        ("DNI:", INGENIERO["dni"]),
        ("Teléfono:", INGENIERO["telefono"]),
        ("Email:", INGENIERO["email"]),
        ("RUC:", INGENIERO["ruc"]),
    ]
    project_data = [
        ("Proyecto:", proyecto or "Por definir"),
        ("Especialidad:", especialidad or "General"),
        ("Fecha:", today.strftime("%d/%m/%Y")),
        ("Escala:", f"1:{escala:.6f}" if escala > 0 else "N/A"),
        ("Documento:", "Informe de Metrados"),
        ("Revisión:", "00 - Preliminar"),
    ]

    row = 3
    for (lbl, val), (r_lbl, r_val) in zip(info_data, project_data):
        _apply_cell(ws.cell(row=row, column=1), lbl, _LABEL_FONT, _INFO_LABEL_FILL, _LEFT_ALIGNMENT)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        _apply_cell(ws.cell(row=row, column=2), val, _VALUE_FONT, alignment=_LEFT_ALIGNMENT)
        _apply_cell(ws.cell(row=row, column=8), r_lbl, _LABEL_FONT, _INFO_LABEL_FILL, _LEFT_ALIGNMENT)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=14)
        _apply_cell(ws.cell(row=row, column=9), r_val, _VALUE_FONT, alignment=_LEFT_ALIGNMENT)
        row += 1

    ws.row_dimensions[row].height = 4
    row += 1

    # Encabezado de tabla
    HEADER_COLS = [
        ("Partida", 14),
        ("Descripción", 40),
        ("Und", 8),
        ("N° veces", 10),
        ("Largo (m)", 10),
        ("Ancho (m)", 10),
        ("Alto (m)", 10),
        ("Parcial", 12),
        ("Total", 12),
        ("Confianza", 10),
        ("Fuente", 16),
        ("Nota", 30),
    ]

    col = 1
    for header_text, width in HEADER_COLS:
        _apply_cell(ws.cell(row=row, column=col), header_text, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGNMENT)
        ws.column_dimensions[get_column_letter(col)].width = width
        col += 1

    ws.row_dimensions[row].height = 22
    return row + 1  # primera fila de datos


def _write_metrado_rows(ws, start_row: int, items: list[MetradoItem]):
    """Escribe las filas de metrado con formato."""
    for i, item in enumerate(items):
        r = start_row + i
        fill = _ROW_EVEN_FILL if i % 2 == 0 else _ROW_ODD_FILL

        _apply_cell(ws.cell(row=r, column=1), item.partida, _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=2), item.descripcion, _DATA_FONT, fill, _LEFT_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=3), item.unidad, _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=9), item.cantidad, _DATA_FONT, fill, _RIGHT_ALIGNMENT,
                    number_format="#,##0.000")
        _apply_cell(ws.cell(row=r, column=10), f"{item.confianza:.0%}", _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=11), item.fuente, _NOTE_FONT, fill, _LEFT_ALIGNMENT)

        # Columnas vacías pero con borde
        for col in [4, 5, 6, 7, 8]:
            _apply_cell(ws.cell(row=r, column=col), None, _DATA_FONT, fill, _CENTER_ALIGNMENT)

        ws.row_dimensions[r].height = 20

    return start_row + len(items)


def _write_resumen_sheet(ws, items: list[MetradoItem], proyecto: str):
    """Crea la hoja de resumen. Recibe el worksheet directamente."""
    row = 1
    ws.merge_cells("A1:F1")
    _apply_cell(ws["A1"], f"RESUMEN DE METRADO - {proyecto}", _TITLE_FONT,
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    row = 3
    resumen_headers = ["Partida", "Descripción", "Und", "Cantidad", "Confianza", "Fuente"]
    for col, h in enumerate(resumen_headers, 1):
        _apply_cell(ws.cell(row=row, column=col), h, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGNMENT)
    ws.row_dimensions[row].height = 22

    for i, item in enumerate(items):
        r = row + 1 + i
        fill = _ROW_EVEN_FILL if i % 2 == 0 else _ROW_ODD_FILL
        _apply_cell(ws.cell(row=r, column=1), item.partida, _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=2), item.descripcion, _DATA_FONT, fill, _LEFT_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=3), item.unidad, _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=4), item.cantidad, _DATA_FONT, fill, _RIGHT_ALIGNMENT,
                    number_format="#,##0.000")
        _apply_cell(ws.cell(row=r, column=5), f"{item.confianza:.0%}", _DATA_FONT, fill, _CENTER_ALIGNMENT)
        _apply_cell(ws.cell(row=r, column=6), item.fuente, _NOTE_FONT, fill, _LEFT_ALIGNMENT)

    # Anchos
    widths = [12, 45, 8, 14, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------

def build_excel(items: list[MetradoItem],
                proyecto: str = "Sin proyecto",
                especialidad: str = "General",
                escala: float = 1.0) -> bytes:
    """Genera un libro Excel profesional con metrados, resumen y diagnóstico.

    Args:
        items: Lista de MetradoItem (agrupado por build_metrado).
        proyecto: Nombre del proyecto/plano.
        especialidad: Especialidad (Arquitectura, Estructuras, etc.).
        escala: Factor de escala usado.

    Returns:
        Bytes del archivo .xlsx.
    """
    wb = Workbook()

    # --- Hoja 1: Metrado ---
    ws_metrado = wb.active
    ws_metrado.title = "Metrado CAPECO"
    data_start = _write_header_block(ws_metrado, proyecto, especialidad, escala=escala)
    _write_metrado_rows(ws_metrado, data_start, items)

    # --- Hoja 2: Resumen ---
    ws_resumen = wb.create_sheet("Resumen")
    _write_resumen_sheet(ws_resumen, items, proyecto)

    # --- Hoja 3: Diagnóstico (info del archivo) ---
    ws_diag = wb.create_sheet("Diagnóstico")
    diag_data = [
        ("Campo", "Valor"),
        ("Proyecto", proyecto),
        ("Especialidad", especialidad),
        ("Fecha", date.today().strftime("%d/%m/%Y")),
        ("Escala", str(escala)),
        ("Total partidas", str(len(items))),
        ("Ingeniero", INGENIERO["nombre"]),
        ("CIP", INGENIERO["cip"]),
        ("Email", INGENIERO["email"]),
    ]
    for i, (campo, valor) in enumerate(diag_data):
        r = i + 1
        fill = _HEADER_FILL if r == 1 else (_ROW_EVEN_FILL if r % 2 == 0 else _ROW_ODD_FILL)
        font = _HEADER_FONT if r == 1 else _DATA_FONT
        _apply_cell(ws_diag.cell(row=r, column=1), campo, font, fill, _LEFT_ALIGNMENT)
        _apply_cell(ws_diag.cell(row=r, column=2), valor, font, fill, _LEFT_ALIGNMENT)
    ws_diag.column_dimensions["A"].width = 20
    ws_diag.column_dimensions["B"].width = 50

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def build_simple_excel(items: list[MetradoItem]) -> bytes:
    """Versión simple para vista previa rápida (pandas)."""
    output = BytesIO()
    data = [it.as_dict() for it in items]
    pd.DataFrame(data).to_excel(output, index=False, sheet_name="Metrado")
    output.seek(0)
    return output.read()
