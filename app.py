from collections import Counter
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
import re

import fitz
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from src.exporters.excel_exporter import build_excel, build_simple_excel
from src.exporters.pdf_exporter import build_pdf
from src.measurements import Measurement, build_metrado, measurements_to_rows, MetradoItem
from src.processors.autocad_processor import extract_dwg_measurements
from src.processors.dxf_processor import extract_dxf_measurements
from src.processors.pdf_processor import extract_pdf_measurements


SUPPORTED_TYPES = ["dwg", "dxf", "pdf"]
APP_DIR = Path(__file__).resolve().parent
CAPECO_TEMPLATE = APP_DIR / "data" / "templates" / "excel" / "capeco_metrado_templates.xlsx"


def save_upload(uploaded_file) -> Path:
    suffix = Path(uploaded_file.name).suffix.lower()
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return Path(tmp.name)


def extract_measurements(file_path: Path, scale_factor: float) -> list[Measurement]:
    suffix = file_path.suffix.lower()
    if suffix == ".dxf":
        return extract_dxf_measurements(file_path, scale_factor=scale_factor)
    if suffix == ".pdf":
        return extract_pdf_measurements(file_path, scale_factor=scale_factor)
    if suffix == ".dwg":
        return extract_dwg_measurements(file_path, scale_factor=scale_factor)
    raise ValueError(f"Formato no soportado: {suffix}")


def analyze_pdf(file_path: Path) -> dict:
    with fitz.open(file_path) as doc:
        page = doc[0]
        text = page.get_text("text") or ""
        drawings = page.get_drawings()
        rect = page.rect

    tags = re.findall(r"\b([PVM]\d+)\*?\b", text)
    unique_tags = sorted(set(tags), key=lambda tag: (tag[0], int(re.search(r"\d+", tag).group())))
    counts = Counter(tag[0] for tag in unique_tags)
    upper = text.upper()

    finishes = {
        "PISO PORCELANATO": upper.count("PISO PORCELANATO"),
        "PISO CEMENTO PULIDO": upper.count("PISO CEMENTO PULIDO"),
        "FCR BALDOSAS YESO": upper.count("FCR BALDOSAS YESO"),
    }
    ambientes_base = [
        "IMAGEN INSTITUCIONAL",
        "SALA DE ESTAR",
        "SECRETARÍA",
        "COMEDOR",
        "DATA CENTER",
        "SS.HH. 1",
        "PASILLO 1",
        "PASADIZO",
        "SS.HH. DISC.",
        "HALL 2",
        "HALL 1",
        "SALA DE REUNIONES",
        "SSHH 2",
        "DUCTO",
        "Rampa",
    ]
    ambientes = [name for name in ambientes_base if name.upper() in upper]

    return {
        "page_width": rect.width,
        "page_height": rect.height,
        "text": text,
        "text_chars": len(text),
        "drawings_count": len(drawings),
        "tags": unique_tags,
        "counts": counts,
        "finishes": finishes,
        "ambientes": ambientes,
    }


def build_capeco_tables(file_name: str, analysis: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    counts = analysis["counts"]
    finishes = analysis["finishes"]

    rows = [
        {
            "Partida Nro.": "OE.3-VANOS-P",
            "Especificaciones": "Puertas identificadas por etiquetas P en plano",
            "Nro. de veces": counts.get("P", 0),
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": counts.get("P", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("P", 0),
            "Und": "und",
            "Nota de revision": "Preliminar. Falta contrastar con cuadro de vanos si existe.",
        },
        {
            "Partida Nro.": "OE.3-VANOS-V",
            "Especificaciones": "Ventanas identificadas por etiquetas V en plano",
            "Nro. de veces": counts.get("V", 0),
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": counts.get("V", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("V", 0),
            "Und": "und",
            "Nota de revision": "Preliminar. Falta contrastar con cuadro de vanos si existe.",
        },
        {
            "Partida Nro.": "OE.3-EQ-M",
            "Especificaciones": "Elementos identificados con etiqueta M",
            "Nro. de veces": counts.get("M", 0),
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": counts.get("M", 0),
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": counts.get("M", 0),
            "Und": "und",
            "Nota de revision": "Requiere leyenda del plano para definir si es mampara, mueble u otro elemento.",
        },
        {
            "Partida Nro.": "OE.3.4.2",
            "Especificaciones": "Piso porcelanato detectado en ambientes del primer nivel",
            "Nro. de veces": finishes["PISO PORCELANATO"],
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None,
            "Und": "m2",
            "Nota de revision": "Cantidad pendiente: falta delimitar areas por ambiente o leer DWG por capas.",
        },
        {
            "Partida Nro.": "OE.3.4.2",
            "Especificaciones": "Piso cemento pulido detectado",
            "Nro. de veces": finishes["PISO CEMENTO PULIDO"],
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None,
            "Und": "m2",
            "Nota de revision": "Cantidad pendiente: falta delimitar area del ambiente correspondiente.",
        },
        {
            "Partida Nro.": "OE.3.3.6",
            "Especificaciones": "Falso cielorraso con baldosas de yeso detectado",
            "Nro. de veces": finishes["FCR BALDOSAS YESO"],
            "Largo": None,
            "Ancho": None,
            "Altura": None,
            "Parcial": None,
            "Vanos o disminuciones - Nro. veces": None,
            "Vanos o disminuciones - Largo": None,
            "Vanos o disminuciones - Ancho": None,
            "Vanos o disminuciones - Altura": None,
            "Vanos o disminuciones - Parcial": None,
            "Total": None,
            "Und": "m2",
            "Nota de revision": "Cantidad pendiente: requiere asociar poligono de cielorraso por ambiente.",
        },
    ]
    metrado_df = pd.DataFrame(rows)

    resumen_df = pd.DataFrame(
        [
            {"Concepto": "Puertas", "Partida": "OE.3-VANOS-P", "Und": "und", "Total": counts.get("P", 0), "Estado": "Preliminar por conteo de etiquetas"},
            {"Concepto": "Ventanas", "Partida": "OE.3-VANOS-V", "Und": "und", "Total": counts.get("V", 0), "Estado": "Preliminar por conteo de etiquetas"},
            {"Concepto": "Elementos M", "Partida": "OE.3-EQ-M", "Und": "und", "Total": counts.get("M", 0), "Estado": "Pendiente de leyenda"},
            {"Concepto": "Piso porcelanato", "Partida": "OE.3.4.2", "Und": "m2", "Total": None, "Estado": "Pendiente de area"},
            {"Concepto": "Piso cemento pulido", "Partida": "OE.3.4.2", "Und": "m2", "Total": None, "Estado": "Pendiente de area"},
            {"Concepto": "Falso cielorraso baldosas yeso", "Partida": "OE.3.3.6", "Und": "m2", "Total": None, "Estado": "Pendiente de area"},
        ]
    )
    diag_df = pd.DataFrame(
        [
            {"Campo": "Archivo", "Valor": file_name},
            {"Campo": "Especialidad estimada", "Valor": "Arquitectura"},
            {"Campo": "Formato de salida", "Valor": "CAPECO General"},
            {"Campo": "Tamano pagina PDF", "Valor": f"{analysis['page_width']:.1f} x {analysis['page_height']:.1f}"},
            {"Campo": "Texto extraido", "Valor": analysis["text_chars"]},
            {"Campo": "Objetos vectoriales", "Valor": analysis["drawings_count"]},
            {"Campo": "Etiquetas detectadas", "Valor": ", ".join(analysis["tags"])},
            {"Campo": "Ambientes detectados", "Valor": ", ".join(analysis["ambientes"])},
        ]
    )
    return metrado_df, resumen_df, diag_df


def build_formatted_capeco_excel(file_name: str, metrado_df: pd.DataFrame, resumen_df: pd.DataFrame) -> bytes:
    wb = load_workbook(CAPECO_TEMPLATE)
    ws = wb["CAPECO General"]
    header_values = {
        "B3": "Magdalena",
        "H3": "1 de 1",
        "B4": "Por definir",
        "H4": file_name,
        "B5": "2026-05-25",
        "H5": "SAS metrados",
        "B6": "Pendiente",
    }
    for cell, value in header_values.items():
        ws[cell] = value

    template_cols = [
        "Partida Nro.",
        "Especificaciones",
        "Nro. de veces",
        "Largo",
        "Ancho",
        "Altura",
        "Parcial",
        "Vanos o disminuciones - Nro. veces",
        "Vanos o disminuciones - Largo",
        "Vanos o disminuciones - Ancho",
        "Vanos o disminuciones - Altura",
        "Vanos o disminuciones - Parcial",
        "Total",
        "Und",
    ]
    for row_idx, row in enumerate(metrado_df[template_cols].itertuples(index=False, name=None), start=11):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = None if pd.isna(value) else value

    rev = wb["CAPECO Revision Parametrica"]
    for cell, value in header_values.items():
        rev[cell] = value
    for row_idx, row in enumerate(resumen_df.itertuples(index=False), start=11):
        rev.cell(row=row_idx, column=1).value = row.Concepto
        rev.cell(row=row_idx, column=2).value = row.Total
        rev.cell(row=row_idx, column=6).value = row.Estado

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


st.set_page_config(page_title="SAS Metrados", layout="wide")

st.markdown(
    """
    <style>
    .block-container {padding-top: 1.2rem;}
    div[data-testid="stMetric"] {background: #f8fafc; border: 1px solid #dbe3ea; padding: 10px; border-radius: 6px;}
    .stButton button {width: 100%; border-radius: 6px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("SAS Metrados")

with st.sidebar:
    st.header("Plano")
    uploaded = st.file_uploader("Subir DWG, DXF o PDF", type=SUPPORTED_TYPES, accept_multiple_files=False)

    st.caption(
        "💡 **Recomendación:** Los archivos **DXF** se procesan más rápido "
        "y **no requieren AutoCAD**. Si tu plano está en DWG, "
        "conviértelo a DXF desde AutoCAD (GUARDARCOMO → .dxf) "
        "o un visor gratuito."
    )

    scale_factor = st.number_input("Factor de escala a metros", min_value=0.000001, value=1.0, step=0.01, format="%.6f")
    specialty = st.selectbox("Especialidad", ["Arquitectura", "Estructuras", "Instalaciones sanitarias", "Instalaciones electricas"])
    output_format = st.selectbox("Formato de tabla", ["CAPECO General", "CAPECO Concreto Armado", "CAPECO Fierro"])
    run_analysis = st.button("Analizar plano", type="primary")
    run_capeco = st.button("Generar tablas CAPECO")

if uploaded is None:
    st.info("Sube un plano para habilitar el diagnostico y la generacion de tablas.")
    st.stop()

file_path = save_upload(uploaded)
st.caption(f"Archivo cargado: `{uploaded.name}`")

if run_analysis or run_capeco:
    st.session_state["last_file"] = str(file_path)
    st.session_state["last_name"] = uploaded.name
    try:
        if file_path.suffix.lower() == ".pdf":
            st.session_state["pdf_analysis"] = analyze_pdf(file_path)
        else:
            st.session_state["measurements"] = extract_measurements(file_path, scale_factor)
    except Exception as exc:
        st.error(f"No se pudo analizar el archivo: {exc}")
        st.stop()

tabs = st.tabs(["Diagnostico", "Tabla CAPECO", "Mediciones base", "Descargas"])

with tabs[0]:
    if "pdf_analysis" in st.session_state:
        analysis = st.session_state["pdf_analysis"]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Texto", analysis["text_chars"])
        c2.metric("Vectores", analysis["drawings_count"])
        c3.metric("Puertas", analysis["counts"].get("P", 0))
        c4.metric("Ventanas", analysis["counts"].get("V", 0))
        st.subheader("Etiquetas detectadas")
        st.write(", ".join(analysis["tags"]) or "Sin etiquetas")
        st.subheader("Ambientes detectados")
        st.write(", ".join(analysis["ambientes"]) or "Sin ambientes detectados")
        st.subheader("Acabados detectados")
        st.dataframe(pd.DataFrame([analysis["finishes"]]), use_container_width=True, hide_index=True)
    else:
        st.info("Presiona `Analizar plano` para ver diagnostico.")

with tabs[1]:
    if "measurements" in st.session_state:
        items = build_metrado(st.session_state["measurements"])
        st.session_state["capeco_metrado_items"] = items
        st.subheader("Metrado generado automáticamente")
        if items:
            df = pd.DataFrame([it.as_dict() for it in items])
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("No se encontraron elementos metrables después de filtrar referencias.")
    elif "pdf_analysis" in st.session_state:
        metrado_df, resumen_df, diag_df = build_capeco_tables(uploaded.name, st.session_state["pdf_analysis"])
        st.session_state["capeco_metrado_df"] = metrado_df
        st.session_state["capeco_resumen_df"] = resumen_df
        st.session_state["capeco_diag_df"] = diag_df
        st.subheader("Metrado CAPECO General")
        st.dataframe(metrado_df, use_container_width=True, hide_index=True)
        st.subheader("Hoja resumen")
        st.dataframe(resumen_df, use_container_width=True, hide_index=True)
    else:
        st.info("Presiona `Analizar plano` para generar la tabla de metrados.")

with tabs[2]:
    if "measurements" in st.session_state:
        raw_df = pd.DataFrame(measurements_to_rows(st.session_state["measurements"]))
        st.dataframe(raw_df, use_container_width=True, hide_index=True)
    elif "pdf_analysis" in st.session_state:
        preview = st.session_state["pdf_analysis"]["text"][:3000]
        st.text_area("Texto extraido del PDF", preview, height=360)
    else:
        st.info("Sin mediciones base todavia.")

with tabs[3]:
    has_metrado = "capeco_metrado_items" in st.session_state
    has_measurements = "measurements" in st.session_state or "pdf_analysis" in st.session_state

    if has_metrado:
        items = st.session_state["capeco_metrado_items"]
        proyecto = uploaded.name

        # Excel profesional
        excel_bytes = build_excel(
            items=items,
            proyecto=proyecto,
            especialidad=specialty,
            escala=scale_factor,
        )
        st.download_button(
            "📥 Descargar Excel profesional (.xlsx)",
            excel_bytes,
            file_name=f"metrado_{Path(uploaded.name).stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # PDF profesional
        pdf_bytes = build_pdf(
            items=items,
            proyecto=proyecto,
            especialidad=specialty,
        )
        st.download_button(
            "📥 Descargar PDF profesional",
            pdf_bytes,
            file_name=f"metrado_{Path(uploaded.name).stem}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

        # Mantener Excel simple como alternativa
        simple_bytes = build_simple_excel(items)
        st.download_button(
            "Descargar Excel simple (vista rápida)",
            simple_bytes,
            file_name=f"metrado_simple_{Path(uploaded.name).stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    elif has_measurements:
        st.info("Presiona `Generar tablas CAPECO` para generar los reportes descargables.")
    else:
        st.info("Sube un plano y genera las tablas para habilitar descargas.")
